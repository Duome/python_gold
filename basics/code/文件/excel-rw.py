# -*- coding: utf-8 -*-
__author__ = 'Duome'

""" 读取excel内容，存入内存中，内存结构为字典
"""

from openpyxl import *
from openpyxl.compat import range


# 获取数据
cont = {}
work_book = load_workbook(filename='anjuke.xlsx')
sheetname = work_book.get_sheet_names()

for i in sheetname:
    cont[i] = []
    sheet = work_book.get_sheet_by_name(i)
    row_nums = len(sheet['A'])
    col_nums = len(sheet['1'])
    for row in range(row_nums):
        cont[i].append([])
        for col in range(col_nums):
            cont[i][row].append(sheet.cell(row=row+1, column=col+1).value)



# 新建excel，写入数据
new_book = Workbook()

worksheet = new_book.active
new_book.remove_sheet(worksheet)

for sheet_name in sheetname:
    worksheets = new_book.create_sheet(title='%s' % sheet_name)
    sheet_cont = cont[sheet_name]
    for row in sheet_cont:
        worksheets.append(row)

new_book.save('new_book.xlsx')


