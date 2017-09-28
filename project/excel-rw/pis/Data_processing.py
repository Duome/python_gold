# -*- coding: utf-8 -*-
__author__ = 'Duome'


from openpyxl import *
from openpyxl.compat import range


# 获取excel数据
class Data_processing(object):

    def get_data(self, filename):
        self.data = {}
        self.filename = filename
        book = load_workbook(filename=self.filename)
        self.sheetnames = book.get_sheet_names()
        for sheetname in self.sheetnames:
            self.data[sheetname] = []
            sheet = book.get_sheet_by_name(sheetname)
            row_nums = len(sheet['A'])
            col_nums = len(sheet['1'])
            for row in range(row_nums):
                self.data[sheetname].append([])
                for col in range(col_nums):
                    self.data[sheetname][row].append(sheet.cell(row=row+1, column=col+1).value)
        return self.data


    def save_data(self, ndata):
        self.ndata = ndata
        new_book = Workbook()
        worksheet = new_book.active
        new_book.remove_sheet(worksheet)
        for sheetname in self.sheetnames:
            worksheets = new_book.create_sheet(title='%s' % sheetname)
            sheet_cont = self.ndata[sheetname]
            for row in sheet_cont:
                worksheets.append(row)

        new_book.save(filename='search.xlsx')
